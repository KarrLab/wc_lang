""" Classes for reading and writing models to/from files.

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2017-07-25
:Copyright: 2016, Karr Lab
:License: MIT
"""

from wc_lang import core
import numpy as np
import openpyxl
import re
import warnings


class Excel(object):
    """ Reads and writes models from Excel workbooks. """

    @staticmethod
    def read(file_name):
        """ Read model from Excel workbook.

        Args:
            file_name (:obj:`str`): path to Excel workbook

        Returns:
            :obj:`wc_lang.core.Model`: model
        """
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", "Discarded range with reserved name", UserWarning)
            wb = openpyxl.load_workbook(filename=file_name)

        # initialize model object
        model = core.Model()

        '''Read details from Excel'''
        # submodels
        ws = wb['Submodels']
        for iRow in range(2, ws.max_row + 1):
            id = ws.cell(row=iRow, column=1).value
            name = ws.cell(row=iRow, column=2).value
            algorithm = ws.cell(row=iRow, column=3).value
            submodel = core.Submodel(id=id, name=name, algorithm=algorithm)
            model.submodels.append(submodel)

        # compartments
        ws = wb['Compartments']
        for iRow in range(2, ws.max_row + 1):
            model.compartments.append(core.Compartment(
                id=ws.cell(row=iRow, column=1).value,
                name=ws.cell(row=iRow, column=2).value,
                initial_volume=float(ws.cell(row=iRow, column=3).value),
                comments=ws.cell(row=iRow, column=4).value,
            ))

        # species
        ws = wb['Species']
        for iRow in range(2, ws.max_row + 1):
            mw_str = ws.cell(row=iRow, column=5).value
            if mw_str:
                mw = float(mw_str)
            else:
                mw = None

            charge_str = ws.cell(row=iRow, column=6).value
            if charge_str:
                charge = float(charge_str)
            else:
                charge = None

            model.species.append(core.Species(
                id=ws.cell(row=iRow, column=1).value,
                name=ws.cell(row=iRow, column=2).value,
                structure=ws.cell(row=iRow, column=3).value,
                empirical_formula=ws.cell(row=iRow, column=4).value,
                molecular_weight=mw or np.NaN,
                charge=charge,
                type=ws.cell(row=iRow, column=7).value,
                concentrations=[
                    core.Concentration(compartment='c', value=float(ws.cell(row=iRow, column=8).value or 0)),
                    core.Concentration(compartment='e', value=float(ws.cell(row=iRow, column=9).value or 0)),
                ],
                cross_refs=[
                    core.CrossReference(
                        source=ws.cell(row=iRow, column=10).value,
                        id=ws.cell(row=iRow, column=11).value,
                    ),
                ],
                comments=ws.cell(row=iRow, column=12).value,
            ))

        # reactions
        ws = wb['Reactions']

        for iRow in range(2, ws.max_row + 1):
            stoichiometry = parse_stoichiometry(ws.cell(row=iRow, column=4).value)

            rate_law_str = ws.cell(row=iRow, column=6).value
            if rate_law_str:
                rate_law = core.RateLaw(rate_law_str)
            else:
                rate_law = None

            model.reactions.append(core.Reaction(
                id=ws.cell(row=iRow, column=1).value,
                name=ws.cell(row=iRow, column=2).value,
                submodel=ws.cell(row=iRow, column=3).value,
                reversible=stoichiometry['reversible'],
                participants=stoichiometry['participants'],
                enzyme=ws.cell(row=iRow, column=5).value,
                rate_law=rate_law,
                vmax=ws.cell(row=iRow, column=7).value,
                km=ws.cell(row=iRow, column=8).value,
                cross_refs=[
                    core.CrossReference(
                        source=ws.cell(row=iRow, column=9).value,
                        id=ws.cell(row=iRow, column=10).value,
                    ),
                ],
                comments=ws.cell(row=iRow, column=11).value,
            ))

        # parameters
        ws = wb['Parameters']
        for iRow in range(2, ws.max_row + 1):
            model.parameters.append(core.Parameter(
                id=ws.cell(row=iRow, column=1).value,
                name=ws.cell(row=iRow, column=2).value,
                submodel=ws.cell(row=iRow, column=3).value,
                value=float(ws.cell(row=iRow, column=4).value),
                units=ws.cell(row=iRow, column=5).value,
                comments=ws.cell(row=iRow, column=6).value,
            ))

        # references
        ws = wb['References']
        for iRow in range(2, ws.max_row + 1):
            model.references.append(core.Reference(
                id=ws.cell(row=iRow, column=1).value,
                name=ws.cell(row=iRow, column=2).value,
                cross_refs=[
                    core.CrossReference(
                        source=ws.cell(row=iRow, column=3).value,
                        id=ws.cell(row=iRow, column=4).value,
                    ),
                ],
                comments=ws.cell(row=iRow, column=5).value,
            ))

        '''deserialize references'''
        undefined_components = []

        # species concentration
        for species in model.species:
            for conc in species.concentrations:
                id = conc.compartment
                obj = model.get_component_by_id(id, 'compartments')
                if id and obj is None:
                    undefined_components.append(id)
                conc.compartment = obj

        # reaction submodel, participant species, participant compartments, enzymes
        for reaction in model.reactions:
            id = reaction.submodel
            obj = model.get_component_by_id(id, 'submodels')
            if id and obj is None:
                undefined_components.append(id)
            reaction.submodel = obj

            for part in reaction.participants:
                id = part.species
                obj = model.get_component_by_id(id, 'species')
                if id and obj is None:
                    undefined_components.append(id)
                part.species = obj

                id = part.compartment
                obj = model.get_component_by_id(id, 'compartments')
                if id and obj is None:
                    undefined_components.append(id)
                part.compartment = obj

                part.calc_id_name()

            id = reaction.enzyme
            obj = model.get_component_by_id(id, 'species')
            if id and obj is None:
                undefined_components.append(id)
            reaction.enzyme = obj

        # parameter submodels
        for param in model.parameters:
            id = param.submodel
            if id:
                obj = model.get_component_by_id(id, 'submodels')
                if obj is None:
                    undefined_components.append(id)
                param.submodel = obj

        if len(undefined_components) > 0:
            undefined_components = list(set(undefined_components))
            undefined_components.sort()
            raise Exception('Undefined components:\n- {}'.format('\n- '.join(undefined_components)))

        ''' Assemble back references'''
        for submodel in model.submodels:
            submodel.reactions = []
            submodel.species = []
            submodel.parameters = []
        for rxn in model.reactions:
            rxn.submodel.reactions.append(rxn)
            for part in rxn.participants:
                rxn.submodel.species.append('{0}[{1}]'.format(part.species.id, part.compartment.id))
            if rxn.enzyme:
                rxn.submodel.species.append('{0}[{1}]'.format(rxn.enzyme.id, 'c'))
            if rxn.rate_law:
                rxn.submodel.species += rxn.rate_law.get_modifiers(model)

        for param in model.parameters:
            if param.submodel:
                param.submodel.parameters.append(param)

        for submodel in model.submodels:
            species_str_list = list(set(submodel.species))
            species_str_list.sort()
            submodel.species = []
            for index, speciesStr in enumerate(species_str_list):
                species_id, comp_id = speciesStr.split('[')
                comp_id = comp_id[0:-1]
                species_comp = core.SpeciesCompartment(
                    species=model.get_component_by_id(species_id, 'species'),
                    compartment=model.get_component_by_id(comp_id, 'compartments'),
                )
                species_comp.calc_id_name()
                submodel.species.append(species_comp)

        '''Return'''
        return model


def parse_stoichiometry(rxn_str):
    """ Parse a string representing the stoichiometry of a reaction into a Python object.

    Args:
        rxn_str (:obj:`str`): string representation of reaction

    Returns:
        :obj:`dict`: dict representation of reaction stoichiometry with two keys: participants and reversible
    """

    # Split stoichiometry in to global compartment, left-hand side, right-hand side, reversibility indictor
    rxn_match = re.match('(?P<compartment>\[([a-z])\]: )?(?P<lhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?) (?P<direction>[<]?)==> (?P<rhs>((\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])? \+ )*(\(\d*\.?\d*([e][-+]?[0-9]+)?\) )?[a-z0-9\-_]+(\[[a-z]\])?)', rxn_str, flags=re.I)
    if rxn_match is None:
        raise Exception('Invalid stoichiometry: {}'.format(rxn_str))

    # Determine reversiblity
    rxn_dict = rxn_match.groupdict()
    reversible = rxn_dict['direction'] == '<'

    # Determine if global compartment for reaction was specified
    if rxn_dict['compartment'] is None:
        global_comp = None
    else:
        global_comp = re.match('\[(?P<compartment>[a-z])\]', rxn_dict['compartment'],
                               flags=re.I).groupdict()['compartment']

    # initialize array of reaction participants
    participants = []

    # Parse left-hand side
    for rxn_part_str in rxn_dict['lhs'].split(' + '):
        rxn_part_dict = re.match(
            '(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxn_part_str, flags=re.I).groupdict()

        species = rxn_part_dict['species']
        compartment = rxn_part_dict['compartment'] or global_comp
        coefficient = float(rxn_part_dict['coefficient'] or 1)

        participants.append(core.ReactionParticipant(
            species=species,
            compartment=compartment,
            coefficient=-coefficient,
        ))

    # Parse right-hand side
    for rxn_part_str in rxn_dict['rhs'].split(' + '):
        rxn_part_dict = re.match(
            '(\((?P<coefficient>\d*\.?\d*([e][-+]?[0-9]+)?)\) )?(?P<species>[a-z0-9\-_]+)(\[(?P<compartment>[a-z])\])?', rxn_part_str, flags=re.I).groupdict()

        species = rxn_part_dict['species']
        compartment = rxn_part_dict['compartment'] or global_comp
        coefficient = float(rxn_part_dict['coefficient'] or 1)

        participants.append(core.ReactionParticipant(
            species=species,
            compartment=compartment,
            coefficient=coefficient,
        ))

    return {
        'reversible': reversible,
        'participants': participants,
    }
