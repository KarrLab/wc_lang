""" Migration to ObjTables format as of 2020-03-09

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2020-03-09
:Copyright: 2020, Karr Lab
:License: MIT
"""

import wc_lang
import openpyxl
import re
import stringcase


def transform(filename):
    # read
    wb = openpyxl.load_workbook(filename=filename)

    for ws in wb:
        if not ws.title.startswith('!'):
            continue

        # lower camel case document and class attributes
        table_head_cell = None
        if isinstance(ws.cell(1, 1).value, str) and ws.cell(1, 1).value.startswith('!!'):
            table_head_cell = ws.cell(1, 1)
            matches = re.findall(r" +(.*?)=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                ws.cell(1, 1).value)
            heading, _, _ = ws.cell(1, 1).value.partition(' ')
            for key, val, _, _ in matches:
                heading += ' {}={}'.format(stringcase.camelcase(key), val)
            ws.cell(1, 1).value = heading

        if isinstance(ws.cell(2, 1).value, str) and ws.cell(2, 1).value.startswith('!!'):
            table_head_cell = ws.cell(2, 1)
            matches = re.findall(r" +(.*?)=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                ws.cell(2, 1).value)
            heading, _, _ = ws.cell(2, 1).value.partition(' ')
            for key, val, _, _ in matches:
                heading += ' {}={}'.format(stringcase.camelcase(key), val)
            ws.cell(2, 1).value = heading

        # set schema
        schema = 'wc_lang'

        if ws.title == '!!_Schema':
            raise NotImplementedError('setting schema name not supported')
        elif ws.title != '!!_Table of contents':
            table_head_cell.value += " schema='{}'".format(schema)

        # set table format
        if table_head_cell:
            if ws.title in ['!!_Schema', '!!_Table of contents']:
                table_head_cell.value += " tableFormat='row'"
            else:
                match = re.search(r" +id=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                    table_head_cell.value)
                if match:
                    table_id = match.group(1)[1:-1]
                    if hasattr(wc_lang, table_id):
                        table = getattr(wc_lang, table_id)
                        table_format = table.Meta.table_format.name
                    else:
                        table_format = 'row'
                    table_head_cell.value += " tableFormat='{}'".format(table_format)

    # save
    wb.save(filename)
