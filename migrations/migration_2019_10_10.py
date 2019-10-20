""" Migration to ObjTables format as of 2019-10-10

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2019-10-10
:Copyright: 2019, Karr Lab
:License: MIT
"""

import obj_tables
import openpyxl


def transform(filename):
    # read
    wb = openpyxl.load_workbook(filename=filename)

    for ws in wb:
        if ws.title.startswith('!') and not ws.title.startswith('!!'):
            ws.title = '!' + ws.title

        if isinstance(ws.cell(1, 1).value, str):
            ws.cell(1, 1).value = ws.cell(1, 1).value \
                .replace('TableType', 'Type') \
                .replace('ModelId', 'Id') \
                .replace('ModelName', 'Name')

    if ('!' + obj_tables.SCHEMA_SHEET_NAME) in wb:
        ws = wb['!' + obj_tables.SCHEMA_SHEET_NAME]
        for i_col in range(ws.max_column):
            cell = ws.cell(2, i_col + 1)
            if cell.value == '!Type':
                i_type_col = i_col
        for i_row in range(ws.max_row - 2):
            cell = ws.cell(i_row + 2 + 1, i_type_col + 1)
            if cell.value == 'Model':
                cell.value = 'Class'

    # adding document heading
    if ('!' + obj_tables.TOC_SHEET_NAME) in wb:
        ws = wb['!' + obj_tables.TOC_SHEET_NAME]
    elif ('!' + obj_tables.SCHEMA_SHEET_NAME) in wb:
        ws = wb['!' + obj_tables.SCHEMA_SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]
    ws.insert_rows(1, amount=1)
    ws.cell(1, 1).value = "!!!ObjTables ObjTablesVersion='{}'".format(
        obj_tables.__version__)

    # save
    wb.save(filename)
