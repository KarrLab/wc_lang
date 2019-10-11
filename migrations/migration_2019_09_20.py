""" Migration from original obj_model format to obj_tables SBtab-like format

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2019-09-20
:Copyright: 2019, Karr Lab
:License: MIT
"""

import obj_tables
import openpyxl
import warnings
import wc_lang.io


def transform(filename):
    # read
    wb = openpyxl.load_workbook(filename=filename)

    # TOC
    if ('!' + obj_tables.TOC_SHEET_NAME) in wb:
        toc_sheet_name = '!' + obj_tables.TOC_SHEET_NAME
    elif obj_tables.TOC_SHEET_NAME in wb:
        toc_sheet_name = obj_tables.TOC_SHEET_NAME
    elif 'Table of contents' in wb:
        toc_sheet_name = 'Table of contents'
    else:
        toc_sheet_name = None

    if toc_sheet_name:
        ws = wb[toc_sheet_name]
        ws.title = '!' + obj_tables.TOC_SHEET_NAME
        first_cell = ws.cell(1, 1).value
        if not (isinstance(first_cell, str) and first_cell.startswith('!!')):
            ws.insert_rows(1, amount=1)
        ws.cell(1, 1).value = "!!ObjTables Type='{}' ObjTablesVersion='{}'".format(
            obj_tables.SCHEMA_TABLE_TYPE, obj_tables.__version__)
        ws.cell(2, 1).value = '!Table'
        ws.cell(2, 2).value = '!Description'
        ws.cell(2, 3).value = '!Number of objects'

    # Rename worksheets
    for ws in wb:
        if not ws.title.startswith('!'):
            ws.title = '!' + ws.title

    # add header lines
    for model in wc_lang.io.Writer.MODELS:
        if model.Meta.table_format == obj_tables.TableFormat.row:
            sheet_name = model.Meta.verbose_name_plural
        else:
            sheet_name = model.Meta.verbose_name
        if ('!' + sheet_name) not in wb:
            warnings.warn('Missing model ' + model.__name__)
            continue
        ws = wb['!' + sheet_name]
        ws.insert_rows(1, amount=1)
        ws.cell(1, 1).value = "!!ObjTables Type='Data' Id='{}' ObjTablesVersion='{}'".format(
            model.__name__, obj_tables.__version__)

        n_head_rows = 1
        for attr in model.Meta.attributes.values():
            if isinstance(attr, obj_tables.RelatedAttribute) and \
                    attr.related_class.Meta.table_format == obj_tables.TableFormat.multiple_cells:
                n_head_rows = 2
                break

        if model.Meta.table_format == obj_tables.TableFormat.row:
            for i_row in range(n_head_rows):
                for i_col in range(ws.max_column):
                    cell = ws.cell(i_row + 2, i_col + 1)
                    if cell.value and isinstance(cell.value, str):
                        cell.value = '!' + cell.value
        else:
            for i_row in range(ws.max_row - 1):
                for i_col in range(n_head_rows):
                    cell = ws.cell(i_row + 2, i_col + 1)
                    if cell.value and isinstance(cell.value, str):
                        cell.value = '!' + cell.value

        for merged_cell_range in ws.merged_cells.ranges:
            merged_cell_range.min_row += 1
            merged_cell_range.max_row += 1

    # save
    wb.save(filename)
