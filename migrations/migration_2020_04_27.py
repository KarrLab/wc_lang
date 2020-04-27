""" Migration to ObjTables format as of 2020-04-27

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2020-04-27
:Copyright: 2020, Karr Lab
:License: MIT
"""

import openpyxl
import re


def transform(filename):
    # read
    wb = openpyxl.load_workbook(filename=filename)

    for ws in wb:
        if not ws.title.startswith('!'):
            continue

        if isinstance(ws.cell(1, 1).value, str) and ws.cell(1, 1).value.startswith('!!'):
            ws.cell(1, 1).value = re.sub(r" +id=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")", r" class=\1",
                                         ws.cell(1, 1).value)

        if isinstance(ws.cell(2, 1).value, str) and ws.cell(2, 1).value.startswith('!!'):
            ws.cell(2, 1).value = re.sub(r" +id=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")", r" class=\1",
                                         ws.cell(2, 1).value)

    # save
    wb.save(filename)
